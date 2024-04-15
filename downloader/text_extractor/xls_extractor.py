import openpyxl

def xls_extractor(file_path):
    text = ""
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and isinstance(cell, str):
                    text += cell + " "
    return text
